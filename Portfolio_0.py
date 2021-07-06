from openpyxl import load_workbook
import os
import datetime
import time
import json
import requests
import telegram
import cryptocompare
from openpyxl.worksheet.table import Table


#-------------------- Environment --------------------------------------------------------------------------
# Set working directory and workbook
os.chdir(r"C:\Users\Nico\Documents\Python\Portfolio")
workbook = load_workbook('Portfolio.xlsx')
worksheet = workbook.active

# Time and Date
now = datetime.datetime.now()
date = (now.strftime("%d/%m/%y"))
current_time = (now.strftime("%X"))

# Get Cryptocurrencies prices in USD using Cryptocompare.com API.
cryptocompare.cryptocompare._set_api_key_parameter('59711b0f290e155ff93aa6e10f183764ea6e167caab9379aadc88928308155c1')

# Column Headers
worksheet['A1'] = 'Asset'
worksheet['B1'] = 'Price (USD)'
worksheet['C1'] = 'Amount'
worksheet['D1'] = 'Balance (USD)'

#-------------------- Functions -----------------------------------------------------------------------------
# Insert Assets
def insert_asset():
    asset_free = str(input('\nInsert an asset: '))
    asset = asset_free
    if check_asset(1,asset) != True:
        write_data(1,asset,2)


# def change_amount(data):
#     for rows in range (1 , worksheet.max_row+1):    
#         if worksheet.cell(row = rows, column = 1).value == data:
#             new_amount = float(input('Write the amount of ' + str(data) + ' that you want to add: '))
#             cells = worksheet.cell(row = rows, column = 3)
#             cells.value = new_amount + float(worksheet.cell(row = rows, column = 3))

# Write assets
def write_data(columns, data, max_row_parameter):
    for rows in range (1 , worksheet.max_row + max_row_parameter):    
        if worksheet.cell(row = rows, column = columns).value == None:
            cells = worksheet.cell(row = rows, column = columns)
            cells.value = data
            print(str(data) + ' added to the portfolio\n')
            write_asset_price(data)
            amount_input(data)
            asset_balance(data)

# Insert another asset or quit
def new_asset():
    print('Do you want to add an asset?')
    option = input('"1" to add a new asset \n"2" to check your portfolio  \n"3" to quit \nOption: ')
    if option == str(1):
        insert_asset()
        new_asset()
    elif option == str(2):
        print_portfolio()
        new_asset()
    elif option == str(3):
        print('Byebye')
    else:
        print('\nYou did not choose a correct option, try again\n')
        new_asset()

# Write asset prices
def write_asset_price(data):
    for rows in range (1 , worksheet.max_row+1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            cells = worksheet.cell(row = rows, column = 2)
            btc_price = cryptocompare.get_price([data],['USD'])
            btc_price_usd = btc_price[data]['USD']
            cells.value = btc_price_usd
            
# User inputs amount of asset
def amount_input(data):
    for rows in range (1 , worksheet.max_row+1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            amount = float(input('Write the amount of ' + str(data) + ' that you own: '))
            cells = worksheet.cell(row = rows, column = 3)
            cells.value = amount

# Individual Asset Balance Calculation
def asset_balance(data):
    for rows in range (1 , worksheet.max_row+1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            price = worksheet.cell(row = rows, column= 2).value
            amount = worksheet.cell(row = rows, column= 3).value
            balance = price * amount
            cells = worksheet.cell(row = rows, column = 4)
            cells.value = balance

# Check if Asset is already in file
def check_asset(columns, data):
    for rows in range (1 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = columns).value == data:
            print(str(data) + ' is already in the Portfolio\n')
            return True
            
# Funtion to print portfolio
def print_portfolio():
    print("\n")
    for i in range (0,105):
        print(end='-')
    print("")   
    for row in worksheet.rows:
        for cell in row:
            filled_space = 20 - len(str(cell.value))
            for i in range (1,filled_space):
                if i == 1:
                    print(cell.value,end=" ")
                if i > 1:
                    print(end=" ")    
        print('')
    for i in range (0,105):
        print(end='-')    
    print("\n")
           
# -------------------- Main Program -----------------------------------------------------------------------------
new_asset()

# Saves the file
workbook.save(filename = "Portfolio.xlsx")     
