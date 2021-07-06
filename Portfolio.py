#-------------------- Environment --------------------------------------------------------------------------
# Importing Libraries
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import datetime
import cryptocompare
import time
from additional_functions import cover, intro, wait

# Set working directory and workbook
current_working_directory = os.getcwd()
os.chdir(current_working_directory)
workbook = load_workbook('Portfolio.xlsx')
workbook.active.title = "Cryptofolio"

# Time and Date
now = datetime.datetime.now()
date = (now.strftime("%d/%m/%y"))
current_time = (now.strftime("%X"))

# Get Cryptocurrencies prices in USD using Cryptocompare.com API.
cryptocompare.cryptocompare._set_api_key_parameter('59711b0f290e155ff93aa6e10f183764ea6e167caab9379aadc88928308155c1')

# Column Headers
workbook.active['A1'] = 'Asset'
workbook.active['B1'] = 'Price (USD)'
workbook.active['C1'] = 'Amount'
workbook.active['D1'] = 'Balance (USD)'
workbook.active['E1'] = 'Date'
workbook.active['F1'] = 'Time'

# Time and Date
now = datetime.datetime.now()
date = (now.strftime("%d/%m/%y"))
current_time = (now.strftime("%X"))

#-------------------- Formating ----------------------------------------------------------------------------
def formatting():
    bold = Font( bold = True)
    for cell in workbook.active["1:1"]:
        cell.font = bold

#-------------------- Functions -----------------------------------------------------------------------------
# Decide what to do
def choice():
    print('')
    time.sleep(0.3)
    print('What do you want to do?')
    time.sleep(0.3)
    print('"1" to add a new asset')
    time.sleep(0.3)
    print('"2" to delete an asset')
    time.sleep(0.3)
    print('"3" to check portfolio')
    time.sleep(0.3)
    print('"4" to quit')
    time.sleep(0.3)
    option = str(input('Option: '))
    if option == str(1):
        insert_asset()
    elif option == str(2):
        delete_asset()
    elif option == str(3):
        time.sleep(1)
        print_portfolio()
    elif option == str(4):
        print('')
        time.sleep(1)
        print('Thank you for using this not so well made Python Portfolio')
        time.sleep(2)
        print('Byebye')
        time.sleep(2)
        print('')
        return
    else:
        print('\nYou did not choose a correct option, try again\n')
    choice()

# Insert Assets
def insert_asset():
    while True:
        asset = str(input('\nInsert asset: ')).upper()
        try:
            cryptocompare.get_price([asset],['USD'])[asset]['USD']
            break
        except TypeError:
            print('The token couldnt be found, try with another one')
    if check_asset(asset) == True: # if the asset already exists
        change_amount(asset)
        asset_balance(asset)
        write_date(asset, date)
        write_time(asset, current_time)
    elif check_asset(asset) == False: # if the asset doesn't exists
        write_data(1,asset)

# Delete Asset
def delete_asset():
    while True:
        asset = str(input('\nInsert the asset you want to delete from this portfolio: ')).upper()
        try:
            cryptocompare.get_price([asset],['USD'])[asset]['USD']
            break
        except TypeError:
            print('')
            print('The token you are writing couldnt be found, try with another one.')
    check = check_asset(asset)
    if check == False:
        print(str(asset) + " is not in the portfolio")
        print('Do you want to add it?')
        option = str(input('If Yes, write "1".\nIf No, write "2"\nOption: '))
        if option == str(1):
            write_data(1,asset)
        elif option == str(2):
            return
        else:
            print('\nYou did not choose a correct option, try again\n')
    elif check == True:
        workbook.create_sheet('Deleted_sheet')
        sheet1 = workbook['Cryptofolio']
        sheet2 = workbook['Deleted_sheet']
        for rows in range (1 , sheet1.max_row + 1):
            if sheet1.cell(row = rows, column = 1).value != asset:
                for i in range (1,sheet1.max_column + 1):
                    sheet2.cell(row = rows, column = i).value = sheet1.cell(row = rows, column = i).value
                for rows in range (1,sheet2.max_row + 1 ):
                    if sheet2.cell(row = rows, column = 1).value == None:
                        sheet2. delete_rows(rows)
        sheet1.title = 'Deleted_sheet1'
        sheet2.title = 'Cryptofolio'
        del workbook['Deleted_sheet1']
        workbook.active = workbook['Cryptofolio']
    print(str (asset) + ' was successfully deleted from the portfolio.')
    formatting()

# Check if Asset is already in file
def check_asset(data):
    bandera = None
    for rows in range (2 , workbook.active.max_row + 1):
        if workbook.active.cell(row = rows, column = 1).value == data:
            print(str(data) + ' is already in the Portfolio\n')
            bandera = True
    if bandera == True:
        return True
    else:
        return False

# Write assets
def write_data(columns, data):
    for rows in range (2 , workbook.active.max_row +  2):
        if workbook.active.cell(row = rows, column = columns).value == None:
            cells = workbook.active.cell(row = rows, column = columns)
            cells.value = data
            print(str(data) + ' added to the portfolio\n')
            write_asset_price(data)
            amount_input(data)
            asset_balance(data)
            write_date(data, date)
            write_time(data, current_time)

# Write asset prices
def write_asset_price(data):
    for rows in range (2 , workbook.active.max_row + 1):
        if workbook.active.cell(row = rows, column = 1).value == data:
            cells = workbook.active.cell(row = rows, column = 2)
            cells.value = cryptocompare.get_price([data],['USD'])[data]['USD']

# User inputs amount of asset
def amount_input(data):
    for rows in range (2 , workbook.active.max_row + 1):
        if workbook.active.cell(row = rows, column = 1).value == data:
            amount = float(input('Write the amount of ' + str(data) + ' that you own: '))
            cells = workbook.active.cell(row = rows, column = 3)
            cells.value = amount

# Individual Asset Balance Calculation
def asset_balance(data):
    for rows in range (2 , workbook.active.max_row + 1):
        if workbook.active.cell(row = rows, column = 1).value == data:
            price = workbook.active.cell(row = rows, column= 2).value
            amount = workbook.active.cell(row = rows, column= 3).value
            balance = price * amount
            cells = workbook.active.cell(row = rows, column = 4)
            cells.value = balance

# To change amount
def change_amount(data):
    print('Do you want to add or substract ' + str(data) + '?')
    option = int(input('If Yes, write "1".\nIf No, write "2"\nOption: '))
    if option == 1:
        for rows in range (2 , workbook.active.max_row+1):
            if workbook.active.cell(row = rows, column = 1).value == data:
                current_amount = workbook.active.cell(row = rows, column = 3).value
                new_amount = float(input('Write the amount of ' + str(data) + ' that you want to add: '))
                cells = workbook.active.cell(row = rows, column = 3)
                cells.value = new_amount + float(current_amount)
    if option == 2:
        return

# Writes date
def write_date(data,date):
    for rows in range (2 , workbook.active.max_row + 1):
        if workbook.active.cell(row = rows, column = 1).value == data:
            workbook.active.cell(row = rows, column = 5).value = date # Date

# Writes time
def write_time(data,time):
    for rows in range (2 , workbook.active.max_row + 1):
        if workbook.active.cell(row = rows, column = 1).value == data:
            workbook.active.cell(row = rows, column = 6).value = time # Time

# Funtion to print portfolio
def print_portfolio():
    print("\n")
    for i in range (0,105):
        print(end='-')
    print("")
    for row in workbook.active.rows:
        for cell in row:
            filled_space = 20 - len(str(cell.value))
            for i in range (1,filled_space):
                if i == 1:
                    print(cell.value,end=" ")
                if i > 1:
                    print(end=" ")
        print('')
    for i in range (0,105):
        print(end='-')
    print("\n")

# -------------------- Main Program -----------------------------------------------------------------------------
cover()
intro()
wait()
formatting()
choice()

# Saves the file
workbook.save(filename = "Portfolio.xlsx")
