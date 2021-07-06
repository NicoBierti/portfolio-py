#-------------------- Environment --------------------------------------------------------------------------
# Importing Libraries
from openpyxl import load_workbook
import os
import datetime
import time
import cryptocompare

# Set working directory and workbook
os.chdir(r"C:\Users\Nico\Documents\Python\Portfolio") 
workbook = load_workbook('Portfolio.xlsx')
worksheet = workbook.active

# Time and Date
now = datetime.datetime.now()
date = (now.strftime("%d/%m/%y"))
current_time = (now.strftime("%X"))

# Get Cryptocurrencies prices in USD using Cryptocompare.com API.
cryptocompare.cryptocompare._set_api_key_parameter('59711b0f290e155ff93aa6e10f183764ea6e167caab9379aadc88928308155c1')

# Column Headers
worksheet['A1'] = 'Asset'
worksheet['B1'] = 'Price (USD)'
worksheet['C1'] = 'Amount'
worksheet['D1'] = 'Balance (USD)'
worksheet['E1'] = 'Date'
worksheet['F1'] = 'Time'

# Time and Date
now = datetime.datetime.now()
date = (now.strftime("%d/%m/%y"))
current_time = (now.strftime("%X"))

#-------------------- Functions -----------------------------------------------------------------------------
# Decide what to do
def choice():
    print('')
    print('What do you want to do?')
    option = str(input('"1" to add a new asset\n"2" to delete an asset\n"3" to check your portfolio\n"4" to quit\nOption: '))
    if option == str(1):
        insert_asset()
    elif option == str(2):
        delete_asset()
    elif option == str(3):
        print_portfolio()
    elif option == str(4):
        print('')
        print('Byebye')
        print('')
        return
    else:
        print('\nYou did not choose a correct option, try again\n')
    choice()

# Insert Assets
def insert_asset():
    while True:
        asset = str(input('\nInsert asset: ')).upper()
        try:
            cryptocompare.get_price([asset],['USD'])[asset]['USD']
            break
        except TypeError:
            print('The token couldnt be found, try with another one')
    if check_asset(asset) == True: # if the asset already exists
        change_amount(asset)
        asset_balance(asset)
        write_date(asset, date)
        write_time(asset, current_time)
    elif check_asset(asset) == False: # if the asset doesn't exists
        write_data(1,asset)

# Delete Asset
def delete_asset():
    while True:
        asset = str(input('\nInsert the asset you want to delete from this portfolio: ')).upper()
        try:
            cryptocompare.get_price([asset],['USD'])[asset]['USD']
            break
        except TypeError:
            print('')
            print('The token you are writing couldnt be found, try with another one.')
    check = check_asset(asset)
    if check == False:
        print(str(asset) + " is not in the portfolio")
        print('Do you want to add it?')
        option = str(input('If Yes, write "1".\nIf No, write "2"\nOption: '))
        if option == str(1):
            write_data(1,asset)
        elif option == str(2):
            return
        else:
            print('\nYou did not choose a correct option, try again\n')
    elif check == True:
        for rows in range (2 , worksheet.max_row + 1):    
            if worksheet.cell(row = rows, column = 1).value == asset:
                worksheet.delete_rows(rows)
                print(str (asset) + ' was successfully deleted from the portfolio.')
    
            
# Check if Asset is already in file
def check_asset(data):
    bandera = None
    for rows in range (2 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            print(str(data) + ' is already in the Portfolio\n')
            bandera = True
    if bandera == True:
        return True
    else:
        return False

# Write assets
def write_data(columns, data):
    for rows in range (2 , worksheet.max_row +  2):    
        if worksheet.cell(row = rows, column = columns).value == None:
            cells = worksheet.cell(row = rows, column = columns)
            cells.value = data
            print(str(data) + ' added to the portfolio\n')
            write_asset_price(data)
            amount_input(data)
            asset_balance(data)
            write_date(data, date)
            write_time(data, current_time)  

# Write asset prices
def write_asset_price(data):
    for rows in range (2 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            cells = worksheet.cell(row = rows, column = 2)
            cells.value = cryptocompare.get_price([data],['USD'])[data]['USD']
            
# User inputs amount of asset
def amount_input(data):
    for rows in range (2 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            amount = float(input('Write the amount of ' + str(data) + ' that you own: '))
            cells = worksheet.cell(row = rows, column = 3)
            cells.value = amount

# Individual Asset Balance Calculation
def asset_balance(data):
    for rows in range (2 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            price = worksheet.cell(row = rows, column= 2).value
            amount = worksheet.cell(row = rows, column= 3).value
            balance = price * amount
            cells = worksheet.cell(row = rows, column = 4)
            cells.value = balance

# To change amount 
def change_amount(data):
    print('Do you want to add or substract ' + str(data) + '?')
    option = int(input('If Yes, write "1".\nIf No, write "2"\nOption: '))
    if option == 1: 
        for rows in range (2 , worksheet.max_row+1):    
            if worksheet.cell(row = rows, column = 1).value == data:
                current_amount = worksheet.cell(row = rows, column = 3).value
                new_amount = float(input('Write the amount of ' + str(data) + ' that you want to add: '))
                cells = worksheet.cell(row = rows, column = 3)
                cells.value = new_amount + float(current_amount)
    if option == 2:
        return

# Writes date
def write_date(data,date):
    for rows in range (2 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            cells = worksheet.cell(row = rows, column = 5) # Date
            cells.value = date
    

# Writes time
def write_time(data,time):
    for rows in range (2 , worksheet.max_row + 1):    
        if worksheet.cell(row = rows, column = 1).value == data:
            cells = worksheet.cell(row = rows, column = 6) # Date
            cells.value = time
           
            
# Funtion to print portfolio
def print_portfolio():
    print("\n")
    for i in range (0,105):
        print(end='-')
    print("")   
    for row in worksheet.rows:
        for cell in row:
            filled_space = 20 - len(str(cell.value))
            for i in range (1,filled_space):
                if i == 1:
                    print(cell.value,end=" ")
                if i > 1:
                    print(end=" ")    
        print('')
    for i in range (0,105):
        print(end='-')    
    print("\n")




# -------------------- Main Program -----------------------------------------------------------------------------
choice()
# Saves the file
workbook.save(filename = "Portfolio.xlsx")     
